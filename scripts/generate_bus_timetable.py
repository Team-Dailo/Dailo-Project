"""
충주 버스 시간표 Excel → JSON 변환 스크립트

사용법:
    python3 scripts/generate_bus_timetable.py

입력:  ~/Downloads/chungju_bus_timetable.xlsx
출력:  frontend/constants/busTimetable.json
"""

import openpyxl, json, re
from pathlib import Path

EXCEL_PATH = Path.home() / 'Downloads' / 'chungju_bus_timetable.xlsx'
OUT_PATH = Path(__file__).parent.parent / 'frontend' / 'constants' / 'busTimetable.json'


def get_time(s):
    times = re.findall(r'\d{1,2}:\d{2}', str(s))
    return times[-1] if times else ''


def get_from(s):
    s = str(s)
    if '/' in s:
        last = s.split('/')[-1]
        return re.sub(r'\d{1,2}:\d{2}.*', '', last).strip()
    return re.sub(r'\s*\d{1,2}:\d{2}.*', '', s).strip() or s.strip()


def parse_standard(ws, label):
    """표준 시트: destination=A, routeNo=B, from=C, time=D, waypoints=E+"""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers = [str(v).strip() if v else '' for v in rows[0]]
    wp_headers = headers[4:]
    trips, cur_dest, cur_route = [], '', ''
    for row in rows[1:]:
        cells = [str(v).strip() if v is not None else '' for v in row]
        if cells[0]: cur_dest = cells[0]
        if cells[1]: cur_route = cells[1]
        if not cells[3]: continue
        wps = [{'label': wp_headers[i], 'value': cells[4 + i]}
               for i in range(len(wp_headers)) if 4 + i < len(cells) and cells[4 + i]]
        trips.append({'destination': cur_dest, 'routeNo': cur_route,
                      'from': cells[2], 'time': cells[3], 'waypoints': wps})
    return {'label': label, 'trips': trips}


def parse_hoam(ws):
    """호암지구방면: routeNo=A, from/time=B(합산), waypoints=C~F, dest=G"""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers = [str(v).strip() if v else '' for v in rows[0]]
    wp_headers = headers[2:6]
    trips, cur_route = [], ''
    for row in rows[1:]:
        cells = [str(v).strip() if v is not None else '' for v in row]
        if cells[0]: cur_route = cells[0]
        t = get_time(cells[1])
        f = get_from(cells[1])
        if not t: continue
        wps = [{'label': wp_headers[i], 'value': cells[2 + i]}
               for i in range(len(wp_headers)) if 2 + i < len(cells) and cells[2 + i]]
        dest_raw = cells[6] if len(cells) > 6 else ''
        dest = re.sub(r'\s*(555|600|666|700|777|999)\s*$', '', dest_raw).strip()
        trips.append({'destination': dest, 'routeNo': cur_route,
                      'from': f, 'time': t, 'waypoints': wps})
    return {'label': '호암지구', 'trips': trips}


def parse_gyotongdae(ws):
    """교통대건국대방면: routeNo=A, direction=B, from=C, time=D(터미널), waypoints=E~H"""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers = [str(v).strip() if v else '' for v in rows[0]]
    wp_headers = headers[4:8]
    trips, cur_route = [], ''
    for row in rows[1:]:
        cells = [str(v).strip() if v is not None else '' for v in row]
        if cells[0]: cur_route = cells[0]
        if not cells[3]: continue
        wps = [{'label': wp_headers[i], 'value': cells[4 + i]}
               for i in range(len(wp_headers)) if 4 + i < len(cells) and cells[4 + i]]
        trips.append({'destination': cells[1], 'routeNo': cur_route,
                      'from': cells[2], 'time': cells[3], 'waypoints': wps})
    return {'label': '교통대·건국대', 'trips': trips}


def parse_circular(ws):
    """시내순환소순환: routeNo=A, dest=B(노선명), from/time=C(합산), waypoints=D~H"""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers = [str(v).strip() if v else '' for v in rows[0]]
    wp_headers = headers[3:8]
    trips, cur_route, cur_dest = [], '', ''
    for row in rows[1:]:
        cells = [str(v).strip() if v is not None else '' for v in row]
        if cells[0]: cur_route = cells[0]
        if cells[1]: cur_dest = cells[1]
        t = get_time(cells[2])
        f = get_from(cells[2])
        if not t: continue
        wps = [{'label': wp_headers[i], 'value': cells[3 + i]}
               for i in range(len(wp_headers)) if 3 + i < len(cells) and cells[3 + i]]
        trips.append({'destination': cur_dest, 'routeNo': cur_route,
                      'from': f, 'time': t, 'waypoints': wps})
    return {'label': '시내순환·소순환', 'trips': trips}


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    standard_sheets = [
        ('팔봉대소원주덕노은', '팔봉·대소원·주덕·노은'),
        ('신니방면',           '신니'),
        ('살미수안보방면',     '살미·수안보'),
        ('양성방면',           '양성'),
        ('금가산척방면',       '금가·산척'),
        ('엄정소태방면',       '엄정·소태'),
        ('동량방면',           '동량'),
        ('중앙탑노은방면',     '중앙탑·노은'),
        ('서충주노선',         '서충주'),
        ('의료원마즈막재방면', '의료원·마즈막재'),
        ('댐선착장방면',       '댐·선착장'),
        ('단월방면',           '단월'),
    ]

    timetable = {}
    for sheet_key, label in standard_sheets:
        timetable[sheet_key] = parse_standard(wb[sheet_key], label)

    timetable['호암지구방면']      = parse_hoam(wb['호암지구방면'])
    timetable['교통대건국대방면']  = parse_gyotongdae(wb['교통대건국대방면'])
    timetable['시내순환소순환']    = parse_circular(wb['시내순환소순환'])

    ws_routes = wb['노선목록']
    routes = [
        {'routeNo': str(r[0]).strip(), 'routeName': str(r[1]).strip() if r[1] else ''}
        for r in ws_routes.iter_rows(min_row=2, values_only=True) if r[0]
    ]

    result = {'routes': routes, 'timetable': timetable}
    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, separators=(',', ':'))

    all_routes = {t['routeNo'] for s in result['timetable'].values() for t in s['trips']}
    print(f'완료: {len(all_routes)}개 노선, {OUT_PATH}')


if __name__ == '__main__':
    main()
